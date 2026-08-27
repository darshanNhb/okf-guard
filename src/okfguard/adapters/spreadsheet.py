"""Adapter for spreadsheets (.xlsx) — hidden sheets, rows, columns, and comments.

Spreadsheets have three distinct hiding mechanisms, each checked
independently:

1. **Hidden sheets** — ``sheet_state`` of ``"hidden"`` or
   ``"veryHidden"`` means the entire sheet's content is invisible to a
   casual user.  All cell values from such sheets go into
   ``hidden_spans``.

2. **Hidden rows and columns** — individual rows or columns can be
   hidden while the sheet itself remains visible.  Cells at the
   intersection of a hidden row or column are treated as hidden
   content.

3. **Cell comments** — comments are only visible on hover and are a
   real vector for smuggled text that most users never see.  Comment
   text goes into ``hidden_spans`` separately from cell values.

4. **White-font-on-white-fill** — for visible cells not already caught
   by rules 1–3, the font colour is compared against the cell fill
   colour.  Only explicit RGB values are compared; theme-indexed colours
   cannot be confidently resolved to RGB without the workbook theme XML
   and are skipped (false negative preferred over false positive).  An
   ``unresolved_theme_colors`` counter in ``source_metadata`` records
   how many cells were skipped for auditability.

**Requires:** ``openpyxl`` — install via
``pip install okf-guard[xlsx]``.
"""

from __future__ import annotations

import io
import os

from okfguard.adapters.base import SourceAdapter, _utc_now_iso
from okfguard.core.models import ExtractedContent

# Colour tolerance per channel.
_COLOUR_TOLERANCE = 10


# ---------------------------------------------------------------------------
# Colour helpers
# ---------------------------------------------------------------------------

def _parse_argb(argb: str | None) -> tuple[int, int, int] | None:
    """Parse an openpyxl ARGB hex string (e.g. ``'FF000000'``) to RGB.

    openpyxl stores colours as 8-character ARGB strings where the first
    two characters are the alpha channel.  Returns ``None`` for absent
    or unparseable values.
    """
    if not argb:
        return None
    s = str(argb)
    if len(s) == 8:
        try:
            return (int(s[2:4], 16), int(s[4:6], 16), int(s[6:8], 16))
        except ValueError:
            return None
    if len(s) == 6:
        try:
            return (int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16))
        except ValueError:
            return None
    return None


def _colours_close(
    c1: tuple[int, int, int],
    c2: tuple[int, int, int],
) -> bool:
    """Return ``True`` if two RGB colours are within tolerance."""
    return all(abs(a - b) <= _COLOUR_TOLERANCE for a, b in zip(c1, c2))


def _get_font_rgb(font_color: object) -> tuple[int, int, int] | None:
    """Extract RGB tuple from an openpyxl font ``Color`` object.

    Returns ``(0, 0, 0)`` (black) when colour is not set (the default
    spreadsheet font colour).  Returns ``None`` for theme-indexed or
    indexed colours that cannot be resolved to RGB.
    """
    if font_color is None:
        return (0, 0, 0)  # Default font colour: black

    color_type = getattr(font_color, "type", None)

    if color_type == "rgb":
        return _parse_argb(getattr(font_color, "rgb", None))

    if color_type is None:
        # Not explicitly set — use default black.
        return (0, 0, 0)

    # Theme or indexed colour — cannot resolve.
    return None


def _get_fill_rgb(fill: object) -> tuple[int, int, int] | None:
    """Extract effective background RGB from an openpyxl cell ``Fill``.

    Returns ``(255, 255, 255)`` (white) for cells with no fill — the
    default spreadsheet background.  Returns ``None`` for theme-indexed
    fills that cannot be resolved to RGB.
    """
    if fill is None:
        return (255, 255, 255)

    pattern_type = getattr(fill, "patternType", None)

    if pattern_type is None or pattern_type == "none":
        return (255, 255, 255)  # No fill → white

    if pattern_type == "solid":
        fg = getattr(fill, "fgColor", None)
        if fg is None:
            return (255, 255, 255)
        fg_type = getattr(fg, "type", None)
        if fg_type == "rgb":
            parsed = _parse_argb(getattr(fg, "rgb", None))
            return parsed if parsed is not None else (255, 255, 255)
        if fg_type is None:
            return (255, 255, 255)
        # Theme or indexed → can't resolve
        return None

    # Other pattern types → treat as white for simplicity
    return (255, 255, 255)


def _has_theme_colour(color_obj: object) -> bool:
    """Return ``True`` if *color_obj* is a theme-indexed colour."""
    if color_obj is None:
        return False
    return getattr(color_obj, "type", None) == "theme"


# ---------------------------------------------------------------------------
# SpreadsheetAdapter
# ---------------------------------------------------------------------------

class SpreadsheetAdapter(SourceAdapter):
    """Extract visible cell values and detect hidden content in .xlsx files.

    **Requires:** ``openpyxl``.
    Install via ``pip install okf-guard[xlsx]``.
    """

    @property
    def format_name(self) -> str:
        """Return ``'xlsx'``."""
        return "xlsx"

    def extract(self, source: str | bytes) -> ExtractedContent:
        """Extract content from an ``.xlsx`` file.

        Args:
            source: A file path (``str``) or raw ``.xlsx`` bytes.

        Returns:
            ``ExtractedContent`` with visible cell values as
            tab-separated text, hidden content from hidden sheets,
            rows, columns, comments, and colour-matched cells in
            ``hidden_spans``, plus metadata including sheet count.

        Raises:
            FileNotFoundError: If *source* is a path that doesn't exist.
            ValueError: If the content cannot be parsed as a valid
                ``.xlsx`` file.
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError(
                "SpreadsheetAdapter requires 'openpyxl'.  "
                "Install it with:  pip install okf-guard[xlsx]"
            ) from exc

        path: str | None = None

        try:
            if isinstance(source, bytes):
                wb = openpyxl.load_workbook(
                    io.BytesIO(source), data_only=True,
                )
            elif isinstance(source, str):
                if not os.path.isfile(source):
                    raise FileNotFoundError(
                        f"XLSX file not found: {source!r}"
                    )
                path = source
                wb = openpyxl.load_workbook(source, data_only=True)
            else:
                raise TypeError(
                    f"SpreadsheetAdapter.extract() expects str or bytes, "
                    f"got {type(source).__name__}"
                )
        except (FileNotFoundError, TypeError, ImportError):
            raise
        except Exception as exc:
            raise ValueError(
                f"Failed to parse XLSX: {exc}"
            ) from exc

        visible_parts: list[str] = []
        hidden_spans: list[str] = []
        sheet_count = len(wb.sheetnames)

        # Auditability counter — see module docstring for rationale.
        unresolved_theme_colors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # --- 1. Hidden sheets ---
            if ws.sheet_state != "visible":
                # All content on this sheet is hidden.
                cells_text: list[str] = []
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cells_text.append(str(cell.value))
                if cells_text:
                    hidden_spans.append(
                        f"[sheet '{sheet_name}' (hidden)] "
                        + " | ".join(cells_text)
                    )
                continue  # Don't process cells individually

            # --- Visible sheet: process each cell ---
            sheet_rows: list[str] = []

            for row in ws.iter_rows():
                row_values: list[str] = []

                for cell in row:
                    cell_ref = cell.coordinate or ""

                    # --- 3. Cell comments → hidden_spans ---
                    if cell.comment is not None:
                        comment_text = cell.comment.text
                        if comment_text and comment_text.strip():
                            hidden_spans.append(
                                f"[sheet '{sheet_name}', cell {cell_ref}"
                                f" (comment)] {comment_text}"
                            )

                    if cell.value is None:
                        row_values.append("")
                        continue

                    cell_str = str(cell.value)

                    # --- 2. Hidden rows / columns ---
                    row_hidden = False
                    col_hidden = False

                    try:
                        rd = ws.row_dimensions.get(cell.row)
                        if rd is not None and rd.hidden:
                            row_hidden = True
                    except (AttributeError, KeyError):
                        pass

                    try:
                        col_letter = get_column_letter(cell.column)
                        cd = ws.column_dimensions.get(col_letter)
                        if cd is not None and cd.hidden:
                            col_hidden = True
                    except (AttributeError, KeyError):
                        pass

                    if row_hidden or col_hidden:
                        mechanism = (
                            "hidden row/column" if row_hidden and col_hidden
                            else "hidden row" if row_hidden
                            else "hidden column"
                        )
                        hidden_spans.append(
                            f"[sheet '{sheet_name}', cell {cell_ref}"
                            f" ({mechanism})] {cell_str}"
                        )
                        continue

                    # --- 4. White-font-on-white-fill ---
                    font_rgb = _get_font_rgb(
                        getattr(cell.font, "color", None)
                    )
                    fill_rgb = _get_fill_rgb(cell.fill)

                    if font_rgb is not None and fill_rgb is not None:
                        if _colours_close(font_rgb, fill_rgb):
                            hidden_spans.append(
                                f"[sheet '{sheet_name}', cell {cell_ref}"
                                f" — font color matches fill] {cell_str}"
                            )
                            continue
                    elif (
                        _has_theme_colour(
                            getattr(cell.font, "color", None)
                        )
                        or _has_theme_colour(
                            getattr(cell.fill, "fgColor", None)
                        )
                    ):
                        # At least one colour is theme-indexed — comparison
                        # skipped, but counted for auditability.
                        unresolved_theme_colors += 1

                    row_values.append(cell_str)

                # Only include rows that have at least one non-empty value.
                if any(v for v in row_values):
                    sheet_rows.append("\t".join(row_values))

            if sheet_rows:
                visible_parts.append(
                    f"## Sheet: {sheet_name}\n"
                    + "\n".join(sheet_rows)
                )

        wb.close()

        metadata: dict[str, str] = {
            "format": self.format_name,
            "extracted_at": _utc_now_iso(),
            "sheet_count": str(sheet_count),
        }
        if path is not None:
            metadata["path"] = path
        if unresolved_theme_colors > 0:
            metadata["unresolved_theme_colors"] = str(
                unresolved_theme_colors
            )

        return ExtractedContent(
            text="\n\n".join(visible_parts),
            hidden_spans=hidden_spans,
            source_metadata=metadata,
        )
